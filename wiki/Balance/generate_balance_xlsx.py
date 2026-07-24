#!/usr/bin/env python3
"""Generate They Will Descend balance workbook (phase-by-phase)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("TheyWillDescend_Balance.xlsx")

BASELINE = 99
WRONG_PENALTY = -1

RESOURCES = ["Villager", "Corn", "Wood", "Stone", "Water", "Obsidian", "Blood"]

BUILDINGS = [
    # id, name, start_built, unlock_phase, build_cost, build_s, inputs, output, prod_s, workers_req, max_workers
    (0, "Home", True, 0, "3 Corn + 2 Wood", 4, "—", "Villager", 3, 0, 0),
    (1, "Well", True, 0, "3 Corn + 3 Wood", 3, "—", "Water", 3, 1, 3),
    (2, "Lumber", True, 0, "4 Corn + 2 Stone", 3, "—", "Wood", 3, 1, 3),
    (3, "Farm", False, 0, "3 Wood + 2 Water", 6, "1 Wood + 1 Water", "Corn", 3, 1, 3),
    (4, "Altar", False, None, "5 Stone + 3 Wood + 2 Obsidian", 6, "1 Villager", "Blood", 3, 1, 1),
    (5, "Obsidian", False, 1, "4 Wood + 4 Stone + 2 Corn", 5, "1 Corn + 1 Wood", "Obsidian", 3, 3, 3),
]

# Target feel: 0 easy → 4 on the edge → 5 breather/dev → 6 hard → 7 very hard
# difficulty 1..5; slack_target = ideal Duration−WallNeed; modifier drafts are NOT in code yet
DIFFICULTY = {
    0: dict(label="легко", score=1, slack="15..25s", intent="обучить DnD/оффер, почти без наказаний"),
    1: dict(label="средне", score=2, slack="10..18s", intent="новый ресурс + лёгкий выбор работ"),
    2: dict(label="сложнее", score=3, slack="5..12s", intent="давление corn; можно −% Farm"),
    3: dict(label="сложно", score=4, slack="3..8s", intent="Obsidian 3 workers — пик дефицита людей"),
    4: dict(label="на пределе", score=5, slack="0..5s", intent="mixed parallel — всё не покрыть"),
    5: dict(label="легко / развитие", score=2, slack="15..30s", intent="передышка: освоить Blood/Altar, накопить"),
    6: dict(label="сложно", score=4, slack="3..8s", intent="twin demands; душить StartTimer для финала"),
    7: dict(label="очень сложно", score=5, slack="0..5s", intent="финал; landing doomsday 5..15 после последнего дара"),
}

# Draft phase production modifiers (% to production SPEED). Negative = slower = harder.
# Scope: All | by Output resource | by Building name. Not implemented in game yet.
MODIFIER_DRAFT = {
    0: [],
    1: [("Stone", -10, "если появится quarry — чуть подрезать, чтобы не фармили вперёд")],
    2: [("Corn", -20, "Harvest Pressure: −20% Farm — главный рычаг без резки оффера")],
    3: [("Obsidian", -15, "дорого по времени + уже 3 workers")],
    4: [("All", -10, "глобальный −10% на пределе; или точечно Wood/Stone/Corn −15")],
    5: [("Blood", +10, "передышка: чуть быстрее Altar, дать освоить жертву"), ("All", +5, "лёгкий буст развития")],
    6: [("Obsidian", -20, ""), ("Blood", -10, "twin — обе ветки медленнее")],
    7: [("All", -15, "финальное сжатие; либо только Blood/Obsidian −25")],
}

PHASES = [
    (0, "Dawn Offering", 60, [("Corn", 6, 8)], [0, 1, 2, 3], "Tutorial pressure"),
    (1, "Stone & Timber", 90, [("Stone", 10, 6), ("Wood", 3, 8)], [5], "NO stone producer yet"),
    (2, "Harvest Pressure", 90, [("Corn", 12, 5)], [], "Farm should be online"),
    (3, "Obsidian Idols", 100, [("Obsidian", 4, 18)], [], "Needs Obsidian building"),
    (4, "Mixed Tribute", 100, [("Wood", 8, 5), ("Stone", 8, 5), ("Corn", 6, 4)], [], "Parallel demand"),
    (5, "First Blood", 110, [("Blood", 3, 25)], [], "Altar must be unlocked"),
    (6, "Twin Demands", 120, [("Obsidian", 3, 20), ("Blood", 2, 30)], [], "Split attention"),
    (7, "Final Propitiation", 120, [("Blood", 5, 22), ("Obsidian", 4, 15), ("Corn", 10, 3)], [], "TARGET remaining 5-15s"),
]

CHEAT_LOADOUTS = {
    0: {"Villager": 1, "built": "Home,Well,Lumber"},
    1: {"Villager": 2, "Corn": 4, "Wood": 4, "built": "Home,Well,Lumber"},
    2: {"Villager": 3, "built": "Home,Well,Lumber"},
    3: {"Villager": 3, "Wood": 6, "Stone": 6, "built": "Home,Well,Lumber"},
    4: {"Villager": 4, "built": "Home,Well,Lumber"},
    5: {"Villager": 5, "built": "Home,Well,Lumber"},
    6: {"Villager": 6, "built": "Home,Well,Lumber"},
    7: {"Villager": 8, "built": "Home,Well,Lumber"},
}

VILLAGER_OFFERS = [
    (1, "start free", "—", 0),
    (2, "2 Water", "2×Well", 6),
    (3, "3 Corn + 2 Water", "Farm+Well", 15),
    (4, "4 Corn + 2 Wood", "Farm+Lumber", 18),
    (5, "2 Stone + 3 Wood + 2 Corn", "??+Lumber+Farm", 24),
    (6, "1 Obsidian + 4 Corn", "Obsidian+Farm", 21),
    (7, "2 Obsidian + 2 Blood?", "late game", 30),
    (8, "3 Blood", "Altar sacrifice tradeoff", 27),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="D6EAF8")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E8F5E9")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN


def header_cells(ws, addrs):
    for addr in addrs:
        ws[addr].fill = HEADER_FILL
        ws[addr].font = HEADER_FONT
        ws[addr].border = THIN


def autosize(ws, min_w=8, max_w=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, min(max_w, len(str(cell.value))))
        ws.column_dimensions[letter].width = max(min_w, length + 2)


def mark_input(cell):
    cell.fill = INPUT_FILL
    cell.border = THIN


def mark_calc(cell):
    cell.fill = CALC_FILL
    cell.border = THIN


def sheet_readme(wb):
    ws = wb.create_sheet("00_Method", 0)
    lines = [
        ("They Will Descend — Balance Workbook", True),
        ("", False),
        ("Кривая сложности (целевой feel)", True),
        ("0 легко → 1 средне → 2 сложнее → 3 сложно → 4 на пределе → 5 легко/развитие → 6 сложно → 7 очень сложно.", False),
        ("Фаза 5 — ПЕРЕДЫШКА: дать освоить Blood/Altar и накопить перед финальной дугой.", False),
        ("", False),
        ("Рычаги сложности (в порядке «трогай сначала»)", True),
        ("A) Offer count / Duration / secondsReward", False),
        ("B) Workers / hire cost / перестановки (дефицит людей)", False),
        ("C) Phase production modifiers (−% скорости) — лист 03b_Modifiers (ещё не в коде)", False),
        ("D) Early unlock зданий «не для пирамиды» — стратегия стока", False),
        ("", False),
        ("Метод: этап за этапом (не все 8 сразу)", True),
        ("1. Выбери фазу → 04_Phase_Lab. Сверься с целевым feel на 02_Phases / 03b.", False),
        ("2. START: карты + люди + Built (Cheat Panel → Jump).", False),
        ("3. Slack = Duration − WallTimeNeed (с учётом modifier на TimeCost).", False),
        ("4. Крути knobs: оффер → люди → modifiers % → rewards.", False),
        ("5. Сыграй. OUTCOME → 05_Carryover.", False),
        ("6. OUTCOME N = START N+1.", False),
        ("", False),
        ("Потом: пары (08_Pair_Curve)", True),
        ("Dip в середине пары; посадка к стыку. Особенно 4→5 (сброс) и 6→7 (финальный зажим).", False),
        ("", False),
        ("Финал (06_Final_Landing)", True),
        ("После последнего дара doomsday = 5..15с. Сейчас Gain=200 — надо резать.", False),
        ("", False),
        ("Цвета: жёлтый=INPUT, зелёный=формула", True),
        ("", False),
        ("Дыры до баланса", True),
        ("• Stone producer отсутствует • Altar не unlock • Modifiers не в коде • Workers не ускоряют • Hire-оффер людей нет", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(i, 1, text)
        cell.font = Font(bold=bold, size=14 if i == 1 else 11)
    ws.column_dimensions["A"].width = 110


def sheet_buildings(wb):
    ws = wb.create_sheet("01_Buildings")
    headers = [
        "Id", "Name", "StartBuilt", "UnlockPhase", "BuildCost", "BuildSec",
        "Inputs", "Output", "ProdSec", "WorkersReq", "MaxWorkers",
        "UnitsPerSec_1w", "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for i, b in enumerate(BUILDINGS):
        bid, name, start, unlock, cost, bs, inputs, output, ps, wr, mw = b
        ups = round(1 / ps, 4) if ps else 0
        note = "NEVER unlocked by phase" if unlock is None else ""
        ws.append([
            bid, name, start, unlock if unlock is not None else "—", cost, bs,
            inputs, output, ps, wr, mw, ups, note,
        ])
        r = 2 + i
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = THIN
            if c in (6, 9, 10, 11):
                mark_input(ws.cell(r, c))
            if c == 12:
                mark_calc(ws.cell(r, c))

    ws.append([])
    ws.append(["GAP", "No Stone producer", "Add Quarry or gatherable Stone", "", "", "", "", "Stone", "", "", "", "", "CRITICAL"])
    ws.append(["GAP", "Altar unlock", "Put buildingId 4 into some phase unlockBuildingIds", "", "", "", "", "Blood", "", "", "", "", "CRITICAL"])
    ws.append([])
    ws.append(["Assumption (planned)", "Extra workers multiply speed: rate = base * workers / WorkersReq"])
    ws.append(["Current code", "Workers are ON/OFF threshold — stacking does NOT speed up yet"])
    autosize(ws)


def sheet_phases(wb):
    ws = wb.create_sheet("02_Phases")
    headers = [
        "Phase", "Title", "Feel", "Diff1to5", "SlackTarget", "DurationSec",
        "OfferResource", "OfferCount", "SecRewardEach", "MaxTimerGain",
        "CumDuration", "CumMaxGain", "UnlockBuildings", "Intent",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    cum_d = 0
    cum_g = 0
    row = 2
    for p in PHASES:
        idx, title, dur, offers, unlocks, _notes = p
        d = DIFFICULTY[idx]
        gain = sum(c * r for _, c, r in offers)
        cum_d += dur
        cum_g += gain
        unlock_str = ",".join(str(u) for u in unlocks) if unlocks else "—"
        for i, (res, count, reward) in enumerate(offers):
            if i == 0:
                ws.append([
                    idx, title, d["label"], d["score"], d["slack"], dur,
                    res, count, reward, gain, cum_d, cum_g, unlock_str, d["intent"],
                ])
            else:
                ws.append(["", "", "", "", "", "", res, count, reward, "", "", "", "", ""])
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row, c)
                cell.border = THIN
                if c in (6, 8, 9):
                    mark_input(cell)
                if c in (10, 11, 12):
                    mark_calc(cell)
            row += 1

    total_dur = sum(p[2] for p in PHASES)
    total_gain = sum(sum(c * r for _, c, r in p[3]) for p in PHASES)
    ws.append([])
    ws.append([
        "TOTAL", "", "", "", "", total_dur, "", "", "", total_gain, "", "", "",
        f"baseline {BASELINE} + maxGain {total_gain} − totalDur {total_dur} = {BASELINE + total_gain - total_dur}",
    ])
    ws.append([])
    ws.append(["Кривая feel:", "легко → средне → сложнее → сложно → ПРЕДЕЛ → передышка → сложно → ОЧЕНЬ сложно"])
    ws.append(["Фаза 5 специально легче — развитие/освоение Blood, чтобы 6–7 били контрастом."])
    autosize(ws)


def sheet_modifiers(wb):
    ws = wb.create_sheet("03b_Modifiers")
    ws["A1"] = "PHASE PRODUCTION MODIFIERS — рычаг сложности (ещё НЕ в коде)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    ws["A3"] = "Смысл"
    ws["B3"] = "Множитель скорости производства на фазе. −20% Corn = Farm крафтит дольше → оффер кукурузы давит без правки count/reward."
    ws["A4"] = "Формула"
    ws["B4"] = "EffectiveProdSec = BaseProdSec / (1 + SpeedPct/100). Пример: 3с и −20% → 3/0.8 = 3.75с."
    ws["A5"] = "Когда трогать"
    ws["B5"] = "После оффера/людей: нужно дожать feel, но не ломать экономику соседних фаз."

    headers = [
        "Phase", "Feel", "Scope (All/Resource/Building)", "SpeedPct",
        "BaseProdSec", "EffectiveProdSec", "Effect on WallTime", "Notes",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(7, c, h)
    style_header_row(ws, 7, len(headers))

    row = 8
    for phase_idx in range(8):
        mods = MODIFIER_DRAFT.get(phase_idx, [])
        feel = DIFFICULTY[phase_idx]["label"]
        if not mods:
            ws.cell(row, 1, phase_idx)
            ws.cell(row, 2, feel)
            ws.cell(row, 3, "—")
            ws.cell(row, 4, 0)
            ws.cell(row, 5, 3)
            ws.cell(row, 6, f"=E{row}/(1+D{row}/100)")
            ws.cell(row, 7, "без модификатора")
            ws.cell(row, 8, "чисто / tutorial")
            for c in range(1, 9):
                ws.cell(row, c).border = THIN
            mark_input(ws.cell(row, 4))
            mark_input(ws.cell(row, 5))
            mark_calc(ws.cell(row, 6))
            row += 1
            continue

        for i, (scope, pct, note) in enumerate(mods):
            ws.cell(row, 1, phase_idx if i == 0 else "")
            ws.cell(row, 2, feel if i == 0 else "")
            ws.cell(row, 3, scope)
            ws.cell(row, 4, pct)
            ws.cell(row, 5, 3)
            ws.cell(row, 6, f"=E{row}/(1+D{row}/100)")
            # wall time factor vs base
            ws.cell(row, 7, f"=IF(D{row}=0,1,1/(1+D{row}/100))")
            ws.cell(row, 8, note)
            for c in range(1, 9):
                ws.cell(row, c).border = THIN
            for c in (3, 4, 5, 8):
                mark_input(ws.cell(row, c))
            mark_calc(ws.cell(row, 6))
            mark_calc(ws.cell(row, 7))
            row += 1

    row += 1
    ws.cell(row, 1, "Матрица рычагов vs feel")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    for c, h in enumerate(["Phase", "Feel", "Offer pressure", "Worker deficit", "Prod modifier", "Early unlock bait"], 1):
        ws.cell(row, c, h)
    style_header_row(ws, row, 6)
    row += 1

    # How hard each lever should push — draft
    lever_plan = [
        (0, "легко", "низкий", "слабый", "0", "Farm unlock заранее"),
        (1, "средне", "средний", "средний", "0..−10", "Obsidian unlock"),
        (2, "сложнее", "высокий corn", "средний", "−20 Corn", "сток water/wood"),
        (3, "сложно", "средний count", "СИЛЬНЫЙ (3w)", "−15 Obsidian", "—"),
        (4, "предел", "тройной оффер", "СИЛЬНЫЙ", "−10 All", "нельзя покрыть всё"),
        (5, "передышка", "низкий/обучающий", "слабее", "+5..+10", "Altar unlock здесь!"),
        (6, "сложно", "двойной", "сильный", "−10..−20", "подготовка к 7"),
        (7, "очень сложно", "высокий + landing", "макс", "−15 All", "тютелька 5..15"),
    ]
    for item in lever_plan:
        for c, v in enumerate(item, 1):
            ws.cell(row, c, v)
            ws.cell(row, c).border = THIN
            if c >= 3:
                mark_input(ws.cell(row, c))
        row += 1

    row += 2
    ws.cell(row, 1, "Связь с 04_Phase_Lab")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 1, "TimeCostEach_effective = TimeCostEach_base × (1 / (1+SpeedPct/100))")
    row += 1
    ws.cell(row, 1, "Пример фазы 2: Corn base 9с chain, −20% → 9/0.8 = 11.25с × 12 count = 135с wall > duration 90 → нужен сток/работники.")
    row += 2
    ws.cell(row, 1, "TODO код: PhaseDefinition.productionModifiers[] { resource|building, speedMul } + apply in ProductionBuilding tick.")
    ws.cell(row, 1).font = Font(italic=True, color="666666")

    autosize(ws, max_w=42)


def sheet_time_cost(wb):
    ws = wb.create_sheet("03_TimeCost")
    ws["A1"] = "Теоретическая цена времени 1 единицы (1 worker, без очередей)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")

    headers = ["Resource", "ProducedBy", "ProdSec", "InputChainExtraSec", "TotalSecPerUnit", "UnitsPerMin", "Editable?"]
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 3, len(headers))

    rows = [
        ("Water", "Well", 3, 0, 3, round(60 / 3, 1), "yes"),
        ("Wood", "Lumber", 3, 0, 3, round(60 / 3, 1), "yes"),
        ("Corn", "Farm", 3, 6, 9, round(60 / 9, 1), "yes — chain Wood+Water if empty"),
        ("Stone", "??? MISSING", 0, 0, 999, 0, "MUST FIX"),
        ("Obsidian", "Obsidian", 3, 12, 15, round(60 / 15, 1), "yes — Corn+Wood rough"),
        ("Blood", "Altar", 3, 0, 3, 20, "yes — + Villager opportunity cost"),
        ("Villager", "Home/Offer", 3, 0, 3, 20, "will become offer-cost"),
    ]
    for i, rdata in enumerate(rows):
        ws.append(list(rdata))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(4 + i, c)
            cell.border = THIN
            if c in (3, 4, 5):
                mark_input(cell)

    ws.append([])
    ws.append(["Как пользоваться"])
    ws.append(["OfferCount × TotalSecPerUnit ≈ минимум wall-time на оффер."])
    ws.append(["Если DurationSec < минимума → фаза непроходима без стока с прошлой фазы."])
    ws.append(["InputChainExtraSec = 0, если инпуты уже на старте (carryover)."])
    ws.append(["План workers: 2 на Well → эффективный ProdSec = 3/2 = 1.5."])
    ws.append(["Phase modifiers (03b): EffectiveProdSec = ProdSec / (1+SpeedPct/100). −20% → ×1.25 к времени."])
    autosize(ws, max_w=40)


def sheet_phase_lab(wb):
    ws = wb.create_sheet("04_Phase_Lab")

    ws["A1"] = "PHASE LAB — балансируй ОДНУ фазу"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A3"] = "ActivePhase (0..7)"
    ws["B3"] = 0
    mark_input(ws["B3"])
    ws["C3"] = "← меняй; цифры из 02_Phases + modifiers из 03b"

    ws["A5"] = "PHASE PARAMS"
    ws["A5"].fill = SECTION_FILL
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "Title"
    ws["B6"] = "Dawn Offering"
    mark_input(ws["B6"])
    ws["A7"] = "TargetFeel"
    ws["B7"] = "легко"
    mark_input(ws["B7"])
    ws["C7"] = "0 легко / 1 средне / 2 сложнее / 3 сложно / 4 предел / 5 передышка / 6 сложно / 7 очень сложно"
    ws["A8"] = "DurationSec"
    ws["B8"] = 60
    mark_input(ws["B8"])
    ws["A9"] = "TimerStart (cheat/carry)"
    ws["B9"] = 99
    mark_input(ws["B9"])
    ws["A10"] = "WrongPenalty"
    ws["B10"] = WRONG_PENALTY
    mark_input(ws["B10"])
    ws["A11"] = "ProdSpeedPct (modifier)"
    ws["B11"] = 0
    mark_input(ws["B11"])
    ws["C11"] = "−20 = на 20% медленнее. EffectiveSec = Base/(1+pct/100)"

    ws["A13"] = "OFFER"
    ws["A13"].fill = SECTION_FILL
    ws["A13"].font = Font(bold=True)
    headers = ["Res", "Count", "SecReward", "Gain", "TimeCostBase", "TimeCostEff", "WallTimeNeed"]
    for i, h in enumerate(headers, 1):
        ws.cell(14, i, h)
    style_header_row(ws, 14, 7)

    for i, (res, cnt, rew, tcost) in enumerate([("Corn", 6, 8, 9), ("", 0, 0, 0), ("", 0, 0, 0)]):
        r = 15 + i
        ws.cell(r, 1, res)
        ws.cell(r, 2, cnt)
        ws.cell(r, 3, rew)
        ws.cell(r, 4, f"=B{r}*C{r}")
        ws.cell(r, 5, tcost)
        ws.cell(r, 6, f"=E{r}/(1+$B$11/100)")
        ws.cell(r, 7, f"=B{r}*F{r}")
        for c in (1, 2, 3, 5):
            mark_input(ws.cell(r, c))
        mark_calc(ws.cell(r, 4))
        mark_calc(ws.cell(r, 6))
        mark_calc(ws.cell(r, 7))

    ws["A18"] = "Totals"
    ws["D18"] = "=SUM(D15:D17)"
    ws["G18"] = "=SUM(G15:G17)"
    mark_calc(ws["D18"])
    mark_calc(ws["G18"])
    ws["A19"] = "Slack (Duration − WallNeed)"
    ws["B19"] = "=B8-G18"
    mark_calc(ws["B19"])
    ws["C19"] = "сверь с SlackTarget для Feel (02_Phases). Modifier жмёт Slack без резки оффера."

    ws["A21"] = "START STATE"
    ws["A21"].fill = SECTION_FILL
    ws["A21"].font = Font(bold=True)
    ws["A22"] = "Resource"
    ws["B22"] = "Count"
    header_cells(ws, ["A22", "B22"])
    for i, res in enumerate(RESOURCES):
        ws.cell(23 + i, 1, res)
        ws.cell(23 + i, 2, 1 if res == "Villager" else 0)
        mark_input(ws.cell(23 + i, 2))

    ws["D21"] = "BUILDINGS"
    ws["D21"].fill = SECTION_FILL
    ws["D21"].font = Font(bold=True)
    ws["D22"] = "Building"
    ws["E22"] = "Built(0/1)"
    ws["F22"] = "Workers"
    header_cells(ws, ["D22", "E22", "F22"])
    for i, b in enumerate(BUILDINGS):
        r = 23 + i
        ws.cell(r, 4, b[1])
        ws.cell(r, 5, 1 if b[2] else 0)
        ws.cell(r, 6, 0)
        mark_input(ws.cell(r, 5))
        mark_input(ws.cell(r, 6))

    ws["A31"] = "PLAYTEST OUTCOME"
    ws["A31"].fill = SECTION_FILL
    ws["A31"].font = Font(bold=True)
    for i, (lab, default) in enumerate([
        ("OfferCompleted?", "yes/no"),
        ("Felt vs TargetFeel?", "match / easier / harder"),
        ("SecondsLeftOnOffer", 0),
        ("TimerAtPhaseEnd", 0),
        ("FeltStress (1-5)", 3),
        ("HadToReassignWorkers?", "yes/no"),
        ("ModifierFelt?", "yes/no/n/a"),
        ("BottleneckResource", ""),
        ("Notes", ""),
    ]):
        ws.cell(32 + i, 1, lab)
        ws.cell(32 + i, 2, default)
        mark_input(ws.cell(32 + i, 2))

    ws["A42"] = "Timer after perfect offer"
    ws["B42"] = "=B9+D18-B8"
    mark_calc(ws["B42"])
    ws["C42"] = "TimerStart + MaxGain − Duration"

    ws["A44"] = "Если Feel не попал"
    ws["A45"] = "Слишком легко"
    ws["B45"] = "↑ count / ↓ reward? нет — лучше −SpeedPct или сильнее дефицит людей"
    ws["A46"] = "Слишком сложно"
    ws["B46"] = "сначала убери −modifier, потом ↓ count / дай сток в Carryover"
    ws["A47"] = "Фаза 5 должна ощущаться ЛЕГЧЕ фазы 4 — иначе сломай контраст перед 6–7"

    autosize(ws, max_w=36)


def sheet_carryover(wb):
    ws = wb.create_sheet("05_Carryover")
    ws["A1"] = "CARRYOVER — OUTCOME N → START N+1"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")
    ws["A2"] = "Заполняй End_* после плейтеста. Start_* фазы N>0 = End_* предыдущей (формула)."
    ws["A3"] = "Потом обнови CheatPanelConfig.phaseLoadouts под реалистичный Start."

    headers = [
        "Phase", "Title",
        "StartTimer", "EndTimer",
        "StartVillagers", "EndVillagers",
        "StartCorn", "EndCorn",
        "StartWood", "EndWood",
        "StartStone", "EndStone",
        "StartWater", "EndWater",
        "StartObsidian", "EndObsidian",
        "StartBlood", "EndBlood",
        "BuildingsSnapshot", "Stress1to5", "OK?",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(5, c, h)
    style_header_row(ws, 5, len(headers))

    # End* column index (1-based) for linking Start* of next phase
    end_cols = {
        "StartTimer": 4,
        "StartVillagers": 6,
        "StartCorn": 8,
        "StartWood": 10,
        "StartStone": 12,
        "StartWater": 14,
        "StartObsidian": 16,
        "StartBlood": 18,
    }

    for i, p in enumerate(PHASES):
        idx, title, *_ = p
        load = CHEAT_LOADOUTS.get(idx, {})
        r = 6 + i
        values = [
            idx, title,
            BASELINE if idx == 0 else None,
            None,
            load.get("Villager", 0), None,
            load.get("Corn", 0), None,
            load.get("Wood", 0), None,
            load.get("Stone", 0), None,
            load.get("Water", 0), None,
            load.get("Obsidian", 0), None,
            load.get("Blood", 0), None,
            load.get("built", ""),
            None,
            None,
        ]
        for c, v in enumerate(values, 1):
            ws.cell(r, c, v)

        for c, name in enumerate(headers, 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            if name.startswith("End") or name in ("Stress1to5", "OK?", "BuildingsSnapshot"):
                mark_input(cell)
            elif name.startswith("Start") and idx == 0:
                mark_input(cell)
            elif name.startswith("Start") and idx > 0 and name in end_cols:
                prev = r - 1
                cell.value = f"={get_column_letter(end_cols[name])}{prev}"
                mark_calc(cell)

    ws["A15"] = "Как работать с чит-панелью"
    ws["A16"] = "1) Jump to phase N с loadout ≈ Start_*"
    ws["A17"] = "2) Сыграй фазу"
    ws["A18"] = "3) Запиши End_* → авто-Start для N+1"
    ws["A19"] = "4) Синхронизируй CheatPanelConfig"
    autosize(ws, max_w=18)
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["S"].width = 28


def sheet_final_landing(wb):
    ws = wb.create_sheet("06_Final_Landing")
    ws["A1"] = "FINAL LANDING — после последнего дара остаток таймера 5..15 сек"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Цель"
    ws["B3"] = "Закрыл финальный оффер → на doomsday timer остаётся 5–15с. Не 0 и не 40+."

    ws["A5"] = "INPUTS"
    ws["A5"].fill = SECTION_FILL
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "Timer at START of phase 7"
    ws["B6"] = 80
    mark_input(ws["B6"])
    ws["C6"] = "из 05_Carryover StartTimer фазы 7"

    ws["A7"] = "Phase 7 DurationSec"
    ws["B7"] = 120
    mark_input(ws["B7"])

    ws["A8"] = "Elapsed when LAST card offered"
    ws["B8"] = 110
    mark_input(ws["B8"])
    ws["C8"] = "секунда фазы, когда сдана последняя карта оффера"

    ws["A10"] = "OFFER (phase 7 current)"
    for i, h in enumerate(["Res", "Count", "SecReward", "TotalGain"], 1):
        ws.cell(11, i, h)
    style_header_row(ws, 11, 4)

    for i, (res, cnt, rew) in enumerate([("Blood", 5, 22), ("Obsidian", 4, 15), ("Corn", 10, 3)]):
        r = 12 + i
        ws.cell(r, 1, res)
        ws.cell(r, 2, cnt)
        ws.cell(r, 3, rew)
        ws.cell(r, 4, f"=B{r}*C{r}")
        for c in (1, 2, 3):
            mark_input(ws.cell(r, c))
        mark_calc(ws.cell(r, 4))

    ws["A15"] = "MaxGain"
    ws["B15"] = "=SUM(D12:D14)"
    mark_calc(ws["B15"])

    ws["A17"] = "MODEL A — landing after last card"
    ws["A17"].font = Font(bold=True)
    ws["A18"] = "Doomsday AFTER last card"
    ws["B18"] = "=B6+B15-B8"
    mark_calc(ws["B18"])
    ws["C18"] = "StartTimer + Gain − Elapsed"
    ws["A19"] = "In window?"
    ws["B19"] = '=IF(AND(B18>=$B$32,B18<=$C$32),"YES","NO")'
    mark_calc(ws["B19"])

    ws["A21"] = "MODEL B — какой StartTimer нужен"
    ws["A21"].font = Font(bold=True)
    ws["A22"] = "Target remaining"
    ws["B22"] = 10
    mark_input(ws["B22"])
    ws["A23"] = "Required StartTimer"
    ws["B23"] = "=B22-B15+B8"
    mark_calc(ws["B23"])
    ws["C23"] = "Target − Gain + Elapsed. Фазы 5–6 должны привести сюда."

    ws["A25"] = "MODEL C — диагностика"
    ws["A25"].font = Font(bold=True)
    ws["A26"] = "Excess above max window"
    ws["B26"] = "=MAX(0,B18-$C$32)"
    mark_calc(ws["B26"])
    ws["A27"] = "Deficit below min window"
    ws["B27"] = "=MAX(0,$B$32-B18)"
    mark_calc(ws["B27"])
    ws["A28"] = "Excess>0 → режь rewards / позже закрытие / ниже StartTimer"
    ws["A29"] = "Deficit>0 → больше reward / раньше оффер / выше StartTimer"

    ws["A32"] = "TARGET WINDOW min"
    ws["B32"] = 5
    mark_input(ws["B32"])
    ws["C32"] = 15
    mark_input(ws["C32"])
    ws["D32"] = "max"

    ws.conditional_formatting.add(
        "B18",
        FormulaRule(formula=["AND(B18>=$B$32,B18<=$C$32)"], fill=GOOD_FILL),
    )
    ws.conditional_formatting.add(
        "B18",
        FormulaRule(formula=["OR(B18<$B$32,B18>$C$32)"], fill=BAD_FILL),
    )

    ws["A34"] = "Сценарии (покрути жёлтые)"
    for i, h in enumerate(["Gain", "Elapsed", "StartT", "Landing", "OK?"], 1):
        ws.cell(35, i, h)
    style_header_row(ws, 35, 5)

    scenarios = [
        (200, 120, 99, "текущий maxGain, старт 99 → слишком жирно"),
        (200, 120, 30, "нужен низкий Start для окна"),
        (200, 100, 40, "закрыл раньше"),
        (70, 115, 55, "урезанный reward — реалистичный коридор"),
        (55, 110, 65, "тютелька ~10"),
        (80, 118, 48, "ещё вариант"),
    ]
    for i, (gain, elapsed, start, note) in enumerate(scenarios):
        r = 36 + i
        ws.cell(r, 1, gain)
        ws.cell(r, 2, elapsed)
        ws.cell(r, 3, start)
        ws.cell(r, 4, f"=C{r}+A{r}-B{r}")
        ws.cell(r, 5, f'=IF(AND(D{r}>=$B$32,D{r}<=$C$32),"YES","NO")')
        for c in (1, 2, 3):
            mark_input(ws.cell(r, c))
        mark_calc(ws.cell(r, 4))
        mark_calc(ws.cell(r, 5))
        ws.cell(r, 6, note)

    ws["A44"] = "Вывод"
    ws["A44"].font = Font(bold=True)
    ws["A45"] = "Сейчас maxGain фазы 7 = 5×22 + 4×15 + 10×3 = 200. При Elapsed≈110–120 окно 5..15 НЕДОСТИЖИМО при любом разумном StartTimer."
    ws["A46"] = "Пример: Target=10, Elapsed=115, Gain=200 → Required Start = 10−200+115 = −75."
    ws["A47"] = "Практичный коридор: MaxGain фазы 7 ≈ 40..80, StartTimer с фазы 6 ≈ 50..90, Elapsed ≈ 100..120."
    ws["A48"] = "Души таймер на фазах 5–6 (меньше rewards / дороже крафт), режь SecReward на фазе 7."

    autosize(ws, max_w=42)


def sheet_villagers(wb):
    ws = wb.create_sheet("07_Villagers")
    ws["A1"] = "VILLAGERS — дефицит + оффер найма (план)"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Сейчас: Home пассивно штампует Villager каждые 3с (workersRequired=0)."
    ws["A4"] = "План: каждый следующий житель = оффер с растущей ценой."

    headers = ["Villager#", "OfferCost (draft)", "TimeGate", "EstWallSec", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(6, c, h)
    style_header_row(ws, 6, len(headers))
    for i, row in enumerate(VILLAGER_OFFERS):
        for c, v in enumerate(row, 1):
            ws.cell(7 + i, c, v)
            ws.cell(7 + i, c).border = THIN
            if c in (2, 4):
                mark_input(ws.cell(7 + i, c))

    ws["A17"] = "Дефицит / перестановка"
    ws["A17"].font = Font(bold=True)
    ws["A18"] = "Людей меньше, чем джобов → таскать / ставить двоих на одну ветку."

    for i, h in enumerate(["Фаза", "TargetTotalVillagers", "ActiveJobSlots", "Deficit", "Design intent"], 1):
        ws.cell(22, i, h)
    style_header_row(ws, 22, 5)

    intents = [
        (0, 1, 2, "tutorial — 1 человек, 2 джобы"),
        (1, 2, 3, "появляется давление"),
        (2, 2, 3, "выбор corn vs infra"),
        (3, 3, 4, "Obsidian хочет 3 workers"),
        (4, 3, 5, "mixed — не покрыть всё"),
        (5, 4, 5, "blood = жертва жителя"),
        (6, 4, 5, "twin demands"),
        (7, 5, 6, "финал — жертвы vs крафт"),
    ]
    for i, (ph, tv, slots, intent) in enumerate(intents):
        r = 23 + i
        ws.cell(r, 1, ph)
        ws.cell(r, 2, tv)
        ws.cell(r, 3, slots)
        ws.cell(r, 4, f"=C{r}-B{r}")
        ws.cell(r, 5, intent)
        for c in (2, 3):
            mark_input(ws.cell(r, c))
        mark_calc(ws.cell(r, 4))

    autosize(ws, max_w=40)


def sheet_pair_curve(wb):
    ws = wb.create_sheet("08_Pair_Curve")
    ws["A1"] = "PAIR CURVE — склейка N + N+1"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Цель"
    ws["B3"] = "Середина пары — непонятно «побеждаю/проигрываю»; к стыку — в коридоре."

    ws["A5"] = "Pair label"
    ws["B5"] = "6+7 (prep final landing)"
    mark_input(ws["B5"])

    for i, h in enumerate(["t (sec into pair)", "Timer", "Event note"], 1):
        ws.cell(7, i, h)
    style_header_row(ws, 7, 3)

    sample = [
        (0, 99, "start phase N"),
        (20, 85, "spend / build"),
        (40, 70, "dip — panic?"),
        (55, 95, "offer dump"),
        (70, 88, "phase N+1"),
        (100, 60, "dip again"),
        (140, 75, "recover"),
        (160, 12, "final pair → land 5..15"),
    ]
    for i, (t, timer, note) in enumerate(sample):
        r = 8 + i
        ws.cell(r, 1, t)
        ws.cell(r, 2, timer)
        ws.cell(r, 3, note)
        mark_input(ws.cell(r, 1))
        mark_input(ws.cell(r, 2))

    chart = LineChart()
    chart.title = "Doomsday timer over pair"
    chart.style = 10
    chart.y_axis.title = "Timer"
    chart.x_axis.title = "t"
    data = Reference(ws, min_col=2, min_row=7, max_row=15)
    cats = Reference(ws, min_col=1, min_row=8, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E5")

    ws["A18"] = "Как крутить"
    ws["A19"] = "1) EndTimer фазы N в коридоре (напр. 60..100)."
    ws["A20"] = "2) В N+1: ранний dip, поздний spike оффером."
    ws["A21"] = "3) Для пары 6+7: End фазы 6 ≈ Required StartTimer из 06_Final_Landing."
    ws["A22"] = "4) Пара 4→5: после ПРЕДЕЛА обязательный сброс (передышка) — иначе 6–7 не читаются как пик."
    ws["A23"] = "5) Modifiers можно включать только на одной фазе пары, чтобы dip/spike читался."
    autosize(ws, max_w=36)


def sheet_checklist(wb):
    ws = wb.create_sheet("09_Checklist")
    ws["A1"] = "Порядок работ"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Done"
    ws["B2"] = "Step"
    header_cells(ws, ["A2", "B2"])

    steps = [
        "Закрыть GAP: Stone producer + Altar unlock + (позже) phase modifiers в коде",
        "Фаза 0 ЛЕГКО: 1 villager, corn — slack ~15–25s",
        "Фаза 1 СРЕДНЕ: новый ресурс; End → Carryover",
        "Фаза 2 СЛОЖНЕЕ: corn pressure + draft −20% Corn modifier",
        "Фаза 3 СЛОЖНО: Obsidian 3 workers — дефицит людей",
        "Фаза 4 НА ПРЕДЕЛЕ: mixed; slack ~0–5; −10% All?",
        "Фаза 5 ПЕРЕДЫШКА/РАЗВИТИЕ: легче чем 4; +% Blood; освоить Altar",
        "Фаза 6 СЛОЖНО: twin; подготовить StartTimer фазы 7",
        "Фаза 7 ОЧЕНЬ СЛОЖНО: 06_Final_Landing = YES (5..15)",
        "Пары 4→5 (сброс) и 6→7 (зажим) — кривая",
        "Обновить CheatPanelConfig loadouts = Carryover Start_*",
        "Прогнать полный ран без Jump",
    ]
    for i, s in enumerate(steps):
        ws.cell(3 + i, 1, False)
        ws.cell(3 + i, 2, s)
        mark_input(ws.cell(3 + i, 1))
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 90


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_buildings(wb)
    sheet_phases(wb)
    sheet_time_cost(wb)
    sheet_modifiers(wb)
    sheet_phase_lab(wb)
    sheet_carryover(wb)
    sheet_final_landing(wb)
    sheet_villagers(wb)
    sheet_pair_curve(wb)
    sheet_checklist(wb)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
